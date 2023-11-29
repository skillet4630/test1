import openpyxl

save_path = 'Python/스타트코딩/업무자동화/03.엑셀자동화/11번가.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path, data_only=True) # data_only=True : 데이터로만 가져오겠다

# data 시트 선택
ws = wb['data']

# 01.모든 셀 데이터 가져오기
# -> 행과 열 개수를 아는 경우
for x in range(1, 9+1): # 최대 행 번호
    for y in range(1, 5+1): # 최대 열 번호
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
    
# 값이 None으로 뜬다면 엑셀파일 저장하고 다시 실행해보기

# -> 행과 열 개수를 모르는 경우
for x in range(1, ws.max_row + 1): # 최대 행 번호
    for y in range(1, ws.max_column + 1): # 최대 열 번호
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
    
# 모든 행 가져오기
for row in ws.iter_rows():
    print(row)

# 2번째 행 부터 가져오기
for row in ws.iter_rows(min_row=2):
    print(row)
    
# 2번째 행 부터 5번쨰 행까지 가져오기
for row in ws.iter_rows(min_row=2, max_row=5):
    print(row)
    
# 2~4행 2~4열 가져오기
for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4):
    for cell in row:
        print(cell.value, end=" ")
    print()
