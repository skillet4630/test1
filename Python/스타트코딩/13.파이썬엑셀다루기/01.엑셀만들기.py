# ▶openpyxl
# 파이썬에세 엑셀을 쉽게 다룰 수 있도록 도와주는 라이브러리
# 설치 : pip install openpyxl
import openpyxl

# 1.엑셀파일 만들기
wb = openpyxl.Workbook()

# 2.엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')

# 3.데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = '1'
ws['B2'] = '오일남'

# 4.엑셀 저장하기
wb.save(r'C:\Users\kjc46\OneDrive\바탕 화면\컴퓨터 공부\Python\스타트코딩\13.파이썬엑셀다루기\참가자_data.xlsx') # r : 로우문자열=r뒤에 문장은 그냥 문자취급해라