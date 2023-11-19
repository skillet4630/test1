import openpyxl

fpath = r'C:\Users\kjc46\OneDrive\바탕 화면\컴퓨터 공부\Python\스타트코딩\13.파이썬엑셀다루기\참가자_data.xlsx'
# 1.엑셀파일 불러오기
wb = openpyxl.load_workbook(r'C:\Users\kjc46\OneDrive\바탕 화면\컴퓨터 공부\Python\스타트코딩\13.파이썬엑셀다루기\참가자_data.xlsx')

# 2.엑셀시트 선택
ws = wb['오징어게임']

# 3.데이터 수정하기
ws['A3'] = 456
ws['B3'] = '성기훈'

# 4.엑셀 저장하기
wb.save(fpath)